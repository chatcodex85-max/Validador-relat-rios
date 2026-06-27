from __future__ import annotations

import argparse
import csv
import re
import csv
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from openpyxl import workbook

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - shown to end users
    load_workbook = None


SALA_ERRO = "Sala"
NUMERO_LOGRADOURO_ERRO = "SN ou S/N"
META_CAPACIDADE_ERRO = "Meta capacidade"
JUNCAO_ERRO = "Bloco sem junção"
BLOCO_NOME_ERRO = "Bloco com nome errado"
NOME_UTF8_ERRO = "Nome fora do padrão UTF-8/ABNT2"
COMPLEMENTO_ERRO = "Complemento errado"
TAG_ESCOLAS_NAO_INDICADAS = "Escolas Não Indicadas"
TAG_FALTA_CADEIRAS = "Falta Cadeiras"
TAG_META_ALCANCADA = "Meta Alcançada"
TAG_CAPACIDADE_EXTRA = "Capacidade extra"


@dataclass(frozen=True)
class ValidationErrorRow:
    tipo_relatorio: str
    id_valor: str
    uf: str
    cidade: str
    nome: str
    campo: str
    meta: str
    capacidade_atual: str
    porcentagem: str
    exibido: str
    erro: str
    tag: str


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]", "", text.lower())


def display(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = display(value).replace("%", "")
    if not text:
        return None
    if "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def format_percentage(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.2f}%"


def meta_capacity_tag(percentage: float | None) -> str:
    if percentage is None:
        return ""
    rounded_percentage = round(percentage, 2)
    if percentage > 100:
        return TAG_CAPACIDADE_EXTRA
    if rounded_percentage == 100:
        return TAG_META_ALCANCADA
    if 90 <= percentage < 100:
        return TAG_FALTA_CADEIRAS
    if 0 <= percentage <= 80:
        return TAG_ESCOLAS_NAO_INDICADAS
    return ""


def first_matching_column(
    headers: dict[str, int], candidates: Iterable[str]
) -> str | None:
    normalized_candidates = {normalize_header(candidate) for candidate in candidates}
    for header in headers:
        if normalize_header(header) in normalized_candidates:
            return header
    return None


def first_id_column(headers: dict[str, int], preferred: Iterable[str]) -> str | None:
    preferred_column = first_matching_column(headers, preferred)
    if preferred_column:
        return preferred_column

    for header in headers:
        if normalize_header(header).startswith("id"):
            return header

    return first_matching_column(headers, ["Ordem", "CodigoIBGE", "IdSincad"])


def first_name_column(headers: dict[str, int], preferred: Iterable[str]) -> str | None:
    preferred_column = first_matching_column(headers, preferred)
    if preferred_column:
        return preferred_column

    for header in headers:
        normalized = normalize_header(header)
        if "nome" in normalized or "descricao" in normalized:
            return header

    return None


def detect_report_type(headers: dict[str, int]) -> str:
    if first_matching_column(headers, ["MetaCapacidade"]) and first_matching_column(
        headers, ["CapacidadeIndicada", "CapacidadeAtual"]
    ):
        return "Meta capacidade"
    if first_matching_column(headers, ["DescricaoIndicacaoJuncao"]) and first_matching_column(headers, ["Blocos"]):
        return "Junções"
    if first_matching_column(
        headers, ["NumeroLogradouro", "Numero Logradouro", "N Logradouro"]
    ):
        return "Locais indicados"
    if first_matching_column(headers, ["Sala", "NomeSala", "Nome Sala"]):
        return "Salas"
    return "Desconhecido"


def is_sn_number(value: object) -> bool:
    text = re.sub(r"\s+", "", display(value).upper())
    return text in {"SN", "S/N"}


def contains_sala(value: object) -> bool:
    return "SALA" in display(value).upper()


def normalized_text(value: object) -> str:
    text = display(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return text.upper()


def invalid_block_names(value: object) -> list[str]:
    allowed_prefixes = ("BLOCO", "PREDIO", "PAVILHAO")
    blocks = [part.strip() for part in display(value).split(",") if part.strip()]
    return [block for block in blocks if not normalized_text(block).startswith(allowed_prefixes)]


def has_encoding_issue(value: object) -> bool:
    text = display(value)
    if not text:
        return False

    mojibake_markers = (
        "�",
        "Ãƒ",
        "Ã‚",
        "Ã¡",
        "ÃÁ",
        "Ã©",
        "Ãª",
        "Ã­",
        "Ã³",
        "Ãº",
        "Ã§",
        "Â°",
        "Âº",
        "â€",
    )
    if any(marker in text for marker in mojibake_markers):
        return True

    allowed = re.compile(r"^[A-Za-zÀ-ÖØ-öø-ÿ0-9\s.,;:!?'\"()/\\&ºª°°\-–]+$")
    return not bool(allowed.fullmatch(text))


def is_wrong_complement(value: object) -> bool:
    text = normalized_text(value)
    if not text:
        return False

    if "PROXIMO" in text:
        return False

    terms = (
        "ESCOLA",
        "PREDIO",
        "BAIRRO",
        "CASA",
        "FACULDADE",
        "CENTRO",
        "SEM COMPLEMENTO",
    )
    if any(term in text for term in terms):
        return True

    return bool(re.search(r"(^|[^A-Z0-9])S\s*/?\s*N([^A-Z0-9]|$)|(^|[^A-Z0-9])SN([^A-Z0-9]|$)", text))


def read_headers(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for index, cell in enumerate(
        next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)), start=1
    ):
        if cell is None:
            continue
        header = str(cell).strip()
        if header:
            headers[header] = index
    return headers


def cell_value(row: tuple[object, ...], one_based_index: int | None) -> object:
    if one_based_index is None or one_based_index < 1 or one_based_index > len(row):
        return ""
    return row[one_based_index - 1]


def build_result(
    *,
    tipo_relatorio: str,
    id_valor: str,
    uf: str,
    cidade: str,
    nome: str,
    exibido: str,
    erro: str,
    campo: str = "",
    meta: str = "",
    capacidade_atual: str = "",
    porcentagem: str = "",
    tag: str = "",
) -> ValidationErrorRow:
    return ValidationErrorRow(
        tipo_relatorio=tipo_relatorio,
        id_valor=id_valor,
        uf=uf,
        cidade=cidade,
        nome=nome,
        campo=campo,
        meta=meta,
        capacidade_atual=capacidade_atual,
        porcentagem=porcentagem,
        exibido=exibido,
        erro=erro,
        tag=tag,
    )


def validate_workbook(path: Path) -> list[ValidationErrorRow]:
    if load_workbook is None:
        raise RuntimeError("Instale a dependencia: pip install -r requirements.txt")

    workbook = load_workbook(path, read_only=True, data_only=True)
    errors: list[ValidationErrorRow] = []

    for sheet in workbook.worksheets:
        headers = read_headers(sheet)
        if not headers:
            continue

        report_type = detect_report_type(headers)
        numero_col = first_matching_column(
            headers, ["NumeroLogradouro", "Numero Logradouro", "N Logradouro"]
        )
        sala_col = first_matching_column(headers, ["Sala", "NomeSala", "Nome Sala"])
        juncao_col = first_matching_column(headers, ["DescricaoIndicacaoJuncao", "Descricao Indicacao Juncao"])
        blocos_col = first_matching_column(headers, ["Blocos"])
        complemento_col = first_matching_column(headers, ["Complemento"])
        nome_utf8_col = first_matching_column(headers, ["DescricaoLocalProva", "Descricao Local Prova", "LocalProva"])
        meta_col = first_matching_column(headers, ["MetaCapacidade", "Meta Capacidade"])
        capacidade_col = first_matching_column(
            headers,
            [
                "CapacidadeIndicada",
                "Capacidade Atual",
                "CapacidadeAtual",
                "CapacidadeIndicacao",
            ],
        )
        taxa_col = first_matching_column(
            headers, ["TaxaIndicacao", "Taxa Indicacao", "Porcentagem"]
        )

        if report_type == "Meta capacidade":
            id_col = first_matching_column(headers, ["CodigoIBGE", "IdSincad"])
        else:
            id_col = first_id_column(headers, ["IdLocalProva", "Id Local Prova"])

        nome_col = first_name_column(
            headers,
            [
                "DescricaoLocalProva",
                "Descricao Local Prova",
                "LocalProva",
                "DescricaoCidade",
            ],
        )
        uf_col = first_matching_column(headers, ["UF", "SiglaUF", "Sigla UF"])
        cidade_col = first_matching_column(
            headers, ["Cidade", "Municipio", "DescricaoCidade"]
        )

        for row in sheet.iter_rows(min_row=2, values_only=True):
            id_valor = display(cell_value(row, headers.get(id_col, 0)))
            uf = display(cell_value(row, headers.get(uf_col, 0)))
            cidade = display(cell_value(row, headers.get(cidade_col, 0)))
            nome = display(cell_value(row, headers.get(nome_col, 0)))

            if report_type == "Meta capacidade" and meta_col and capacidade_col:
                meta_value = cell_value(row, headers[meta_col])
                capacidade_value = cell_value(row, headers[capacidade_col])
                taxa_value = cell_value(row, headers.get(taxa_col, 0))
                meta = parse_number(meta_value)
                capacidade = parse_number(capacidade_value)
                percentage = parse_number(taxa_value)

                if percentage is None and meta and capacidade is not None:
                    percentage = (capacidade / meta) * 100

                tag = meta_capacity_tag(percentage)
                if tag:
                    errors.append(
                        build_result(
                            tipo_relatorio=report_type,
                            id_valor=id_valor,
                            uf=uf,
                            cidade=cidade,
                            nome=nome or cidade,
                            meta=display(meta_value),
                            capacidade_atual=display(capacidade_value),
                            porcentagem=format_percentage(percentage),
                            exibido=display(taxa_value)
                            or format_percentage(percentage),
                            erro=META_CAPACIDADE_ERRO,
                            campo=taxa_col or "TaxaIndicacao",
                            tag=tag,
                        )
                    )

            if numero_col:
                numero_value = cell_value(row, headers[numero_col])
                if is_sn_number(numero_value):
                    errors.append(
                        build_result(
                            tipo_relatorio=report_type,
                            id_valor=id_valor,
                            uf=uf,
                            cidade=cidade,
                            nome=nome,
                            exibido=display(numero_value),
                            erro=NUMERO_LOGRADOURO_ERRO,
                            campo=numero_col,
                        )
                    )

            if sala_col:
                sala_value = cell_value(row, headers[sala_col])
                if contains_sala(sala_value):
                    sala_nome = display(sala_value)
                    errors.append(
                        build_result(
                            tipo_relatorio=report_type,
                            id_valor=id_valor,
                            uf=uf,
                            cidade=cidade,
                            nome=nome or sala_nome,
                            exibido=sala_nome,
                            erro=SALA_ERRO,
                            campo=sala_col,
                        )
                    )

            if juncao_col:
                juncao_value = display(cell_value(row, headers[juncao_col]))
                if not juncao_value:
                    errors.append(
                        build_result(
                            tipo_relatorio=report_type,
                            id_valor=id_valor,
                            uf=uf,
                            cidade=cidade,
                            nome=nome,
                            exibido="(vazio)",
                            erro=JUNCAO_ERRO,
                            campo=juncao_col,
                        )
                    )

            if blocos_col:
                wrong_blocks = invalid_block_names(cell_value(row, headers[blocos_col]))
                if wrong_blocks:
                    errors.append(
                        build_result(
                            tipo_relatorio=report_type,
                            id_valor=id_valor,
                            uf=uf,
                            cidade=cidade,
                            nome=nome,
                            exibido=", ".join(wrong_blocks),
                            erro=BLOCO_NOME_ERRO,
                            campo=blocos_col,
                        )
                    )

            if nome_utf8_col:
                nome_utf8_value = cell_value(row, headers[nome_utf8_col])
                if has_encoding_issue(nome_utf8_value):
                    errors.append(
                        build_result(
                            tipo_relatorio=report_type,
                            id_valor=id_valor,
                            uf=uf,
                            cidade=cidade,
                            nome=nome or display(nome_utf8_value),
                            exibido=display(nome_utf8_value),
                            erro=NOME_UTF8_ERRO,
                            campo=nome_utf8_col,
                        )
                    )

            if complemento_col:
                complemento_value = cell_value(row, headers[complemento_col])
                if is_wrong_complement(complemento_value):
                    errors.append(
                        build_result(
                            tipo_relatorio=report_type,
                            id_valor=id_valor,
                            uf=uf,
                            cidade=cidade,
                            nome=nome,
                            exibido=display(complemento_value),
                            erro=COMPLEMENTO_ERRO,
                            campo=complemento_col,
                        )
                    )

    workbook.close()
    return errors


def validate_files(paths: Iterable[Path]) -> list[ValidationErrorRow]:
    errors: list[ValidationErrorRow] = []
    for path in paths:
        if not path.exists():
            raise FileNotFoundError(f"Arquivo nao encontrado: {path}")
        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"O arquivo precisa ser .xlsx: {path}")
        errors.extend(validate_workbook(path))
    return errors


def run_cli(args: argparse.Namespace) -> int:
    errors = validate_files([Path(file) for file in args.files])

    if not errors:
        print("Nenhum erro encontrado.")
        return 0

    if all(error.tipo_relatorio == "Meta capacidade" for error in errors):
        print("ID;UF;Cidade;Nome;Meta;Capacidade Atual;Porcentagem;Tag")
        for error in errors:
            print(
                ";".join(
                    [
                        error.id_valor,
                        error.uf,
                        error.cidade,
                        error.nome,
                        error.meta,
                        error.capacidade_atual,
                        error.porcentagem,
                        error.tag,
                    ]
                )
            )
        return 1

    print("ID;UF;Cidade;Nome;Campo;Exibido;Erro;Tag")
    for error in errors:
        print(
            ";".join(
                [
                    error.id_valor,
                    error.uf,
                    error.cidade,
                    error.nome,
                    error.campo,
                    error.exibido,
                    error.erro,
                    error.tag,
                ]
            )
        )
    return 1


def export_to_csv(errors: list[ValidationErrorRow], output_path: Path) -> None:
    columns = columns_for_export(errors)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file, delimiter=";")

        writer.writerow(columns)

        for error in errors:
            writer.writerow(build_export_row(error, columns))


def export_to_xlsx(errors: list[ValidationErrorRow], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resultado"

    columns = columns_for_export(errors)

    sheet.append(columns)

    for error in errors:
        sheet.append(build_export_row(error, columns))

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = max_length + 3

    workbook.save(output_path)


def columns_for_export(errors: list[ValidationErrorRow]) -> list[str]:

    if errors and all(e.tipo_relatorio == "Meta capacidade" for e in errors):
        return [
            "ID",
            "UF",
            "Cidade",
            "Nome",
            "Meta",
            "Capacidade Atual",
            "Porcentagem",
            "Tag",
        ]

    return [
        "ID",
        "UF",
        "Cidade",
        "Nome",
        "Campo",
        "Exibido",
        "Erro",
        "Tag",
    ]


def build_export_row(error: ValidationErrorRow, columns: list[str]) -> list[str]:

    values = {
        "ID": error.id_valor,
        "UF": error.uf,
        "Cidade": error.cidade,
        "Nome": error.nome,
        "Campo": error.campo,
        "Meta": error.meta,
        "Capacidade Atual": error.capacidade_atual,
        "Porcentagem": error.porcentagem,
        "Exibido": error.exibido,
        "Erro": error.erro,
        "Tag": error.tag,
    }

    return [values[column] for column in columns]


def run_gui() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    selected_files: list[Path] = []
    all_errors: list[ValidationErrorRow] = []
    current_columns: list[str] = []

    root = tk.Tk()
    # Forca o registro do root como "default root" do tkinter. Em builds
    # congelados (PyInstaller --onefile --windowed), bibliotecas importadas
    # antes deste ponto podem criar e descartar um interpretador Tcl/Tk
    # proprio, deixando tkinter._default_root como None mesmo depois de
    # tk.Tk() ser chamado. Isso fazia tk.BooleanVar/tk.StringVar sem master
    # explicito falharem com "Too early to create variable: no default root
    # window". Forcar aqui garante que o root atual seja sempre o default.
    tk._default_root = root

    root.title("Validador de relatorios XLSX")
    root.geometry("1420x720")
    dark_theme = tk.BooleanVar(master=root, value=False)

    style = ttk.Style(root)

    column_labels = {
        "id": "ID",
        "uf": "UF",
        "cidade": "Cidade",
        "nome": "Nome",
        "campo": "Campo",
        "meta": "Meta",
        "capacidade_atual": "Capacidade atual",
        "porcentagem": "Porcentagem",
        "exibido": "Exibido",
        "erro": "Erro",
        "tag": "Tag",
    }
    column_widths = {
        "id": 88,
        "uf": 50,
        "cidade": 300,
        "nome": 300,
        "campo": 190,
        "meta": 90,
        "capacidade_atual": 130,
        "porcentagem": 120,
        "exibido": 130,
        "erro": 140,
        "tag": 190,
    }
    column_anchors = {
        "id": tk.W,
        "uf": tk.W,
        "cidade": tk.W,
        "nome": tk.W,
        "campo": tk.W,
        "meta": tk.E,
        "capacidade_atual": tk.E,
        "porcentagem": tk.E,
        "exibido": tk.W,
        "erro": tk.W,
        "tag": tk.W,
    }

    def value_for_column(error: ValidationErrorRow, column: str) -> str:
        values = {
            "id": error.id_valor,
            "uf": error.uf,
            "cidade": error.cidade,
            "nome": error.nome,
            "campo": error.campo,
            "meta": error.meta,
            "capacidade_atual": error.capacidade_atual,
            "porcentagem": error.porcentagem,
            "exibido": error.exibido,
            "erro": error.erro,
            "tag": error.tag,
        }
        return values[column]

    def columns_for_errors(errors: list[ValidationErrorRow]) -> list[str]:
        if errors and all(
            error.tipo_relatorio == "Meta capacidade" for error in errors
        ):
            return [
                "id",
                "uf",
                "cidade",
                "nome",
                "meta",
                "capacidade_atual",
                "porcentagem",
                "tag",
            ]

        columns = ["id", "uf", "cidade", "nome"]
        optional_columns = [
            "campo",
            "meta",
            "capacidade_atual",
            "porcentagem",
            "exibido",
            "erro",
            "tag",
        ]
        for column in optional_columns:
            if any(value_for_column(error, column) for error in errors):
                columns.append(column)
        return columns

    def configure_table_columns(columns: list[str]) -> None:
        current_columns.clear()
        current_columns.extend(columns)
        table["columns"] = columns
        for column in columns:
            table.heading(column, text=column_labels[column])
            table.column(
                column,
                width=column_widths[column],
                minwidth=max(45, column_widths[column] - 30),
                anchor=column_anchors[column],
                stretch=column in {"cidade", "nome", "tag"},
            )
            
    def export_csv():
        if not all_errors:
            messagebox.showwarning("Exportação", "Nenhum resultado para exportar.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
        )
        if filename:
            export_to_csv(filtered_errors(), Path(filename))
            messagebox.showinfo("Sucesso", "Arquivo CSV exportado.")

    def export_xlsx():
        if not all_errors:
            messagebox.showwarning("Exportação", "Nenhum resultado para exportar.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if filename:
            export_to_xlsx(filtered_errors(), Path(filename))
            messagebox.showinfo("Sucesso", "Arquivo XLSX exportado.")

    def apply_theme() -> None:
        if dark_theme.get():
            bg = "#0f1720"
            panel = "#15202b"
            fg = "#e8eef5"
            field = "#101820"
            selected = "#0b78d0"
            button_text = "Tema claro"
        else:
            bg = "#f3f5f7"
            panel = "#ffffff"
            fg = "#111827"
            field = "#ffffff"
            selected = "#0078d7"
            button_text = "Tema escuro"

        root.configure(bg=bg)
        style.theme_use("clam")
        style.configure(".", background=bg, foreground=fg)
        style.configure("TFrame", background=bg)
        style.configure("TLabelframe", background=bg, foreground=fg)
        style.configure("TLabelframe.Label", background=bg, foreground=fg)
        style.configure("TLabel", background=bg, foreground=fg)
        style.configure("TButton", padding=(10, 5))
        style.configure("TEntry", fieldbackground=field, foreground=fg)
        style.configure(
            "TCombobox", fieldbackground=field, background=panel, foreground=fg
        )
        style.configure(
            "Treeview",
            background=panel,
            fieldbackground=panel,
            foreground=fg,
            rowheight=24,
            borderwidth=0,
        )
        style.configure(
            "Treeview.Heading",
            background=field,
            foreground=fg,
            font=("Segoe UI", 9, "bold"),
        )
        style.map(
            "Treeview",
            background=[("selected", selected)],
            foreground=[("selected", "#ffffff")],
        )
        theme_button.config(text=button_text)

    top_frame = ttk.Frame(root, padding=12)
    top_frame.pack(fill=tk.X)

    file_label = ttk.Label(top_frame, text="Nenhum arquivo selecionado")
    file_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

    def refresh_file_label() -> None:
        if selected_files:
            file_label.config(text=" | ".join(path.name for path in selected_files))
        else:
            file_label.config(text="Nenhum arquivo selecionado")

    def choose_files() -> None:
        paths = filedialog.askopenfilenames(
            title="Selecione os relatorios .xlsx",
            filetypes=[("Planilhas Excel", "*.xlsx")],
        )
        if paths:
            selected_files.clear()
            selected_files.extend(Path(path) for path in paths)
            refresh_file_label()

    filter_frame = ttk.LabelFrame(root, text="Filtros", padding=12)
    filter_frame.pack(fill=tk.X, padx=12, pady=(0, 12))

    search_var = tk.StringVar(master=root)
    uf_var = tk.StringVar(master=root, value="Todas")
    cidade_var = tk.StringVar(master=root, value="Todas")
    erro_var = tk.StringVar(master=root, value="Todos")
    tag_var = tk.StringVar(master=root, value="Todas")

    ttk.Label(filter_frame, text="Buscar").grid(row=0, column=0, sticky=tk.W)
    search_entry = ttk.Entry(filter_frame, textvariable=search_var, width=24)
    search_entry.grid(row=1, column=0, padx=(0, 10), sticky=tk.EW)

    ttk.Label(filter_frame, text="UF").grid(row=0, column=1, sticky=tk.W)
    uf_combo = ttk.Combobox(
        filter_frame, textvariable=uf_var, state="readonly", width=10, values=["Todas"]
    )
    uf_combo.grid(row=1, column=1, padx=(0, 10), sticky=tk.EW)

    ttk.Label(filter_frame, text="Cidade").grid(row=0, column=2, sticky=tk.W)
    cidade_combo = ttk.Combobox(
        filter_frame,
        textvariable=cidade_var,
        state="readonly",
        width=22,
        values=["Todas"],
    )
    cidade_combo.grid(row=1, column=2, padx=(0, 10), sticky=tk.EW)

    ttk.Label(filter_frame, text="Erro").grid(row=0, column=3, sticky=tk.W)
    erro_combo = ttk.Combobox(
        filter_frame,
        textvariable=erro_var,
        state="readonly",
        width=18,
        values=["Todos"],
    )
    erro_combo.grid(row=1, column=3, padx=(0, 10), sticky=tk.EW)

    ttk.Label(filter_frame, text="Tag").grid(row=0, column=4, sticky=tk.W)
    tag_combo = ttk.Combobox(
        filter_frame, textvariable=tag_var, state="readonly", width=22, values=["Todas"]
    )
    tag_combo.grid(row=1, column=4, padx=(0, 10), sticky=tk.EW)

    filter_frame.columnconfigure(0, weight=2)
    filter_frame.columnconfigure(2, weight=1)

    def clear_table() -> None:
        for item in table.get_children():
            table.delete(item)

    def filtered_errors() -> list[ValidationErrorRow]:
        search = search_var.get().strip().lower()
        selected_uf = uf_var.get()
        selected_cidade = cidade_var.get()
        selected_erro = erro_var.get()
        selected_tag = tag_var.get()
        result: list[ValidationErrorRow] = []

        for error in all_errors:
            if selected_uf != "Todas" and error.uf != selected_uf:
                continue
            if selected_cidade != "Todas" and error.cidade != selected_cidade:
                continue
            if selected_erro != "Todos" and error.erro != selected_erro:
                continue
            if selected_tag != "Todas" and error.tag != selected_tag:
                continue
            if search:
                searchable = " ".join(
                    [
                        error.tipo_relatorio,
                        error.id_valor,
                        error.uf,
                        error.cidade,
                        error.nome,
                        error.meta,
                        error.capacidade_atual,
                        error.porcentagem,
                        error.exibido,
                        error.erro,
                        error.tag,
                    ]
                ).lower()
                if search not in searchable:
                    continue
            result.append(error)

        return result

    def render_table() -> None:
        errors = filtered_errors()
        clear_table()
        configure_table_columns(columns_for_errors(errors))

        for error in errors:
            table.insert(
                "",
                tk.END,
                values=tuple(
                    value_for_column(error, column) for column in current_columns
                ),
            )

        status.config(
            text=f"{len(errors)} item(ns) exibido(s) de {len(all_errors)} encontrado(s). Clique no ID para copiar."
        )

    def update_filter_options(errors: list[ValidationErrorRow]) -> None:
        uf_combo["values"] = ["Todas"] + sorted(
            {error.uf for error in errors if error.uf}
        )
        cidade_combo["values"] = ["Todas"] + sorted(
            {error.cidade for error in errors if error.cidade}
        )
        erro_combo["values"] = ["Todos"] + sorted(
            {error.erro for error in errors if error.erro}
        )
        tag_combo["values"] = ["Todas"] + sorted(
            {error.tag for error in errors if error.tag}
        )
        search_var.set("")
        uf_var.set("Todas")
        cidade_var.set("Todas")
        erro_var.set("Todos")
        tag_var.set("Todas")

    def validate_selected() -> None:
        if not selected_files:
            messagebox.showwarning("Arquivos", "Selecione pelo menos um arquivo .xlsx.")
            return

        try:
            errors = validate_files(selected_files)
        except Exception as exc:  # pragma: no cover - user feedback
            messagebox.showerror("Erro ao validar", str(exc))
            return

        all_errors.clear()
        all_errors.extend(errors)
        update_filter_options(errors)
        render_table()

    def clear_filters() -> None:
        search_var.set("")
        uf_var.set("Todas")
        cidade_var.set("Todas")
        erro_var.set("Todos")
        tag_var.set("Todas")
        render_table()

    def copy_id_on_click(event) -> None:
        if table.identify_column(event.x) != "#1":
            return

        row_id = table.identify_row(event.y)
        if not row_id:
            return

        id_value = table.item(row_id, "values")[0]
        root.clipboard_clear()
        root.clipboard_append(id_value)
        status.config(text=f"ID {id_value} copiado.")

    def toggle_theme() -> None:
        dark_theme.set(not dark_theme.get())
        apply_theme()

    ttk.Button(top_frame, text="Selecionar .xlsx", command=choose_files).pack(
        side=tk.RIGHT, padx=(8, 0)
    )
    ttk.Button(top_frame, text="Validar", command=validate_selected).pack(
        side=tk.RIGHT, padx=(8, 0)
    )
    theme_button = ttk.Button(top_frame, text="Tema escuro", command=toggle_theme)
    theme_button.pack(side=tk.RIGHT, padx=(8, 0))
    ttk.Button(filter_frame, text="Filtrar", command=render_table).grid(
        row=1, column=5, padx=(0, 8)
    )
    ttk.Button(filter_frame, text="Limpar filtros", command=clear_filters).grid(
        row=1, column=6
    )

    for combo in (uf_combo, cidade_combo, erro_combo, tag_combo):
        combo.bind("<<ComboboxSelected>>", lambda _event: render_table())
    search_entry.bind("<Return>", lambda _event: render_table())

    table_frame = ttk.Frame(root, padding=(12, 0, 12, 8))
    table_frame.pack(fill=tk.BOTH, expand=True)

    table = ttk.Treeview(table_frame, columns=(), show="headings")
    configure_table_columns(["id", "uf", "cidade", "nome", "exibido", "erro"])
    table.bind("<ButtonRelease-1>", copy_id_on_click)

    y_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=table.yview)
    x_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=table.xview)
    table.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
    table.grid(row=0, column=0, sticky="nsew")
    y_scroll.grid(row=0, column=1, sticky="ns")
    x_scroll.grid(row=1, column=0, sticky="ew")
    table_frame.columnconfigure(0, weight=1)
    table_frame.rowconfigure(0, weight=1)

    status = ttk.Label(root, text="Pronto.", padding=(12, 0, 12, 12))
    status.pack(fill=tk.X)
    apply_theme()

    root.mainloop()


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validador de relatorios .xlsx")
    parser.add_argument("files", nargs="*", help="Arquivos .xlsx para validar")
    parser.add_argument("--gui", action="store_true", help="Abrir interface grafica")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    if args.gui or not args.files:
        run_gui()
        return 0
    return run_cli(args)


if __name__ == "__main__":
    raise SystemExit(main())
